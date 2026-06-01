import csv
import io
from datetime import date, datetime, time

from django.core.files.uploadedfile import UploadedFile
from openpyxl import load_workbook
from rest_framework.exceptions import ValidationError as DRFValidationError

from apps.core.models import CatalogItem
from apps.core.serializers import (
    AcademicPeriodSerializer,
    CatalogItemSerializer,
    TimeSlotSerializer,
    WorkingDaySerializer,
)


class MasterDataImportError(Exception):
    pass


RESOURCE_DEFINITIONS = {
    "periods": {
        "headers": ["code", "name", "start_date", "end_date", "is_active"],
        "required_headers": ["code", "name", "start_date", "end_date"],
    },
    "working_days": {
        "headers": ["day_of_week", "name", "is_active"],
        "required_headers": ["day_of_week", "name"],
    },
    "time_slots": {
        "headers": ["name", "start_time", "end_time", "is_active"],
        "required_headers": ["name", "start_time", "end_time"],
    },
    "teacher_link_type": {
        "headers": ["name", "description", "is_active"],
        "required_headers": ["name"],
    },
    "class_type": {
        "headers": ["name", "description", "is_active"],
        "required_headers": ["name"],
    },
    "academic_space_type": {
        "headers": ["name", "description", "is_active"],
        "required_headers": ["name"],
    },
}


def get_import_templates():
    templates = []
    for resource_type, definition in RESOURCE_DEFINITIONS.items():
        templates.append(
            {
                "resource_type": resource_type,
                "headers": definition["headers"],
                "required_headers": definition["required_headers"],
            }
        )
    return templates


def import_master_data(*, file_obj: UploadedFile, resource_type: str):
    if resource_type not in RESOURCE_DEFINITIONS:
        raise MasterDataImportError("El tipo de recurso para importacion no es valido.")

    headers, rows = _parse_rows(file_obj)
    _validate_required_headers(headers, resource_type)

    if not rows:
        return {
            "resource_type": resource_type,
            "total_processed": 0,
            "successful": 0,
            "failed": 0,
            "rows": [],
        }

    normalized_rows = [_normalize_row_keys(row) for row in rows]

    report_rows = []
    successful = 0

    for row_index, raw_row in normalized_rows:
        if _is_blank_row(raw_row):
            continue

        try:
            payload = _build_payload(resource_type, raw_row)
            _save_row(resource_type, payload)
            successful += 1
            report_rows.append(
                {
                    "row": row_index,
                    "status": "success",
                    "message": "Registro importado correctamente.",
                }
            )
        except Exception as exc:  # noqa: BLE001
            report_rows.append(
                {
                    "row": row_index,
                    "status": "error",
                    "message": _format_exception_message(exc),
                }
            )

    total_processed = len(report_rows)
    failed = total_processed - successful

    return {
        "resource_type": resource_type,
        "total_processed": total_processed,
        "successful": successful,
        "failed": failed,
        "rows": report_rows,
    }


def _parse_rows(file_obj: UploadedFile):
    file_name = (file_obj.name or "").lower()
    if file_name.endswith(".csv"):
        return _parse_csv_rows(file_obj)
    if file_name.endswith(".xlsx"):
        return _parse_xlsx_rows(file_obj)

    raise MasterDataImportError("Solo se permiten archivos CSV y XLSX.")


def _parse_csv_rows(file_obj: UploadedFile):
    content = file_obj.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    headers = [str(header or "").strip().lower() for header in (reader.fieldnames or [])]
    rows = list(enumerate(reader, start=2))
    return headers, rows


def _parse_xlsx_rows(file_obj: UploadedFile):
    workbook = load_workbook(filename=io.BytesIO(file_obj.read()), data_only=True)
    worksheet = workbook.active

    values = list(worksheet.iter_rows(values_only=True))
    if not values:
        return [], []

    headers = [str(header or "").strip() for header in values[0]]
    rows = []

    for row_number, row_values in enumerate(values[1:], start=2):
        row_dict = {}
        for column_index, header in enumerate(headers):
            if not header:
                continue
            row_dict[header] = row_values[column_index] if column_index < len(row_values) else None
        rows.append((row_number, row_dict))

    return [header.lower() for header in headers if header], rows


def _normalize_row_keys(row_with_index):
    row_index, row = row_with_index
    row_data = {}
    for key, value in (row or {}).items():
        normalized_key = str(key or "").strip().lower()
        if not normalized_key:
            continue
        row_data[normalized_key] = value
    return row_index, row_data


def _validate_required_headers(headers, resource_type):
    normalized_headers = {header.strip().lower() for header in headers}
    required_headers = set(RESOURCE_DEFINITIONS[resource_type]["required_headers"])

    missing_headers = sorted(required_headers - normalized_headers)
    if missing_headers:
        missing = ", ".join(missing_headers)
        raise MasterDataImportError(f"Faltan columnas requeridas en la plantilla: {missing}.")


def _is_blank_row(raw_row):
    for value in raw_row.values():
        if str(value or "").strip():
            return False
    return True


def _to_text(value):
    return str(value or "").strip()


def _to_bool(value, *, default=True):
    if value is None or str(value).strip() == "":
        return default

    normalized = str(value).strip().lower()
    if normalized in {"true", "1", "si", "sí", "yes", "y"}:
        return True
    if normalized in {"false", "0", "no", "n"}:
        return False

    raise MasterDataImportError("El valor de is_active no es valido.")


def _to_int(value, *, field_name):
    normalized = _to_text(value)
    if not normalized:
        raise MasterDataImportError(f"El campo {field_name} es obligatorio.")
    try:
        return int(normalized)
    except ValueError as exc:
        raise MasterDataImportError(f"El campo {field_name} debe ser numerico.") from exc


def _to_date(value, *, field_name):
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()

    normalized = _to_text(value)
    if not normalized:
        raise MasterDataImportError(f"El campo {field_name} es obligatorio.")

    try:
        return datetime.strptime(normalized, "%Y-%m-%d").date()
    except ValueError as exc:
        raise MasterDataImportError(
            f"El campo {field_name} debe tener formato YYYY-MM-DD."
        ) from exc


def _to_time(value, *, field_name):
    if isinstance(value, time):
        return value

    normalized = _to_text(value)
    if not normalized:
        raise MasterDataImportError(f"El campo {field_name} es obligatorio.")

    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(normalized, fmt).time()
        except ValueError:
            continue

    raise MasterDataImportError(f"El campo {field_name} debe tener formato HH:MM o HH:MM:SS.")


def _build_payload(resource_type, raw_row):
    if resource_type == "periods":
        return {
            "code": _to_text(raw_row.get("code")),
            "name": _to_text(raw_row.get("name")),
            "start_date": _to_date(raw_row.get("start_date"), field_name="start_date"),
            "end_date": _to_date(raw_row.get("end_date"), field_name="end_date"),
            "is_active": _to_bool(raw_row.get("is_active"), default=True),
        }

    if resource_type == "working_days":
        return {
            "day_of_week": _to_int(raw_row.get("day_of_week"), field_name="day_of_week"),
            "name": _to_text(raw_row.get("name")),
            "is_active": _to_bool(raw_row.get("is_active"), default=True),
        }

    if resource_type == "time_slots":
        return {
            "name": _to_text(raw_row.get("name")),
            "start_time": _to_time(raw_row.get("start_time"), field_name="start_time"),
            "end_time": _to_time(raw_row.get("end_time"), field_name="end_time"),
            "is_active": _to_bool(raw_row.get("is_active"), default=True),
        }

    if resource_type in {
        "teacher_link_type",
        "class_type",
        "academic_space_type",
    }:
        return {
            "name": _to_text(raw_row.get("name")),
            "description": _to_text(raw_row.get("description")),
            "is_active": _to_bool(raw_row.get("is_active"), default=True),
        }

    raise MasterDataImportError("El tipo de recurso para importacion no es valido.")


def _save_row(resource_type, payload):
    if resource_type == "periods":
        serializer = AcademicPeriodSerializer(data=payload)
    elif resource_type == "working_days":
        serializer = WorkingDaySerializer(data=payload)
    elif resource_type == "time_slots":
        serializer = TimeSlotSerializer(data=payload)
    else:
        serializer = CatalogItemSerializer(data=payload, context={"catalog_type": resource_type})

    serializer.is_valid(raise_exception=True)
    serializer.save()


def _format_exception_message(exc):
    if isinstance(exc, DRFValidationError):
        payload = exc.detail
        if isinstance(payload, dict):
            first_key = next(iter(payload), None)
            if first_key is not None:
                value = payload[first_key]
                if isinstance(value, list) and value:
                    return str(value[0])
                return str(value)
        if isinstance(payload, list) and payload:
            return str(payload[0])
        return str(payload)

    return str(exc)
